import djclick as click
import numpy as np

from openpyxl import Workbook
from openpyxl import worksheet
from openpyxl.worksheet.table import Table
from typing import Any
from typing import Iterable

from django.contrib.auth import get_user_model
from django.db.models import Count
from django.db.models import Q
from django.db.models import QuerySet

from libraries.models import CommitAuthor

# Set a the number of percentiles to calculate.
# e.g. 20 means calculate every 5 percent
NUMBER_OF_PERCENTILES = 20
STEP = int(100 / NUMBER_OF_PERCENTILES)
PERCENTILE_RANKS = list(range(STEP, (NUMBER_OF_PERCENTILES + 1) * STEP, STEP))


def calculate_percentile(
    data_list: Iterable[int],
):
    a = np.array(data_list)
    res = np.percentile(a, PERCENTILE_RANKS)
    return res


def insert_data_into_column(
    column_label: str,
    data_list: Iterable[Any],
    start_index: int,
    worksheet: worksheet,
):
    index = start_index
    for i in data_list:
        worksheet[f"{column_label}{index}"] = i
        index += 1


def insert_percentile_data_into_column(
    column_label: str,
    data_list: Iterable[int],
    start_index: int,
    worksheet: worksheet,
):
    percentiles = calculate_percentile(data_list)
    insert_data_into_column(column_label, percentiles, start_index, worksheet)


def insert_qs_data_into_percentile_column(
    column_label: str,
    qs: QuerySet,
    start_index: int,
    worksheet: worksheet,
    field_name: str,
):
    data_list = qs.order_by(field_name).values_list(field_name, flat=True)
    title = " ".join(field_name.split("_")).capitalize()
    worksheet[f"{column_label}{start_index}"] = title
    insert_percentile_data_into_column(
        column_label, data_list, start_index + 1, worksheet
    )


@click.command()
def command():
    User = get_user_model()

    user_qs = (
        User.objects.all()
        .annotate(
            library_versions_authored=Count("author_libraryversions", distinct=True),
            libraries_authored=Count("authors", distinct=True),
        )
        .order_by("-email")
        .exclude(Q(library_versions_authored=0) & Q(libraries_authored=0))
    )
    commit_author_qs = (
        CommitAuthor.objects.all()
        .annotate(
            commits_authored=Count("commit", distinct=True),
            reviews_submitted=Count("submitted_reviews", distinct=True),
            mailing_list_count=Count("emaildata", distinct=True),
        )
        .exclude(
            Q(commits_authored=0) & Q(reviews_submitted=0) & Q(mailing_list_count=0)
        )
    )

    wb = Workbook()
    ws = wb.active

    # Using the default sheet, Present Library Version Authors and Library Authors
    ws.title = "Library Authors and Versions"
    ws.append(
        [
            "Email",
            "Libraries Authored",
            "Library Versions Authored",
        ]
    )
    for user in user_qs:
        ws.append(
            [
                user.email,
                user.libraries_authored,
                user.library_versions_authored,
            ]
        )
    tab = Table(
        displayName="LibraryAuthorsandVersions", ref=f"A1:C{user_qs.count() + 1}"
    )
    ws.add_table(tab)

    wb.create_sheet("Commit Author Data")
    ws = wb["Commit Author Data"]
    ws.append(
        [
            "Name",
            "Commits Authored",
            "Reviews Submitted",
            "Mailing List Contributions",
        ]
    )
    for user in commit_author_qs:
        ws.append(
            [
                user.display_name,
                user.commits_authored,
                user.reviews_submitted,
                user.mailing_list_count,
            ]
        )
    tab = Table(
        displayName="CommitAuthorData", ref=f"A1:D{commit_author_qs.count() + 1}"
    )
    ws.add_table(tab)

    # Add percentile data
    wb.create_sheet("Percentile Analysis")
    ws = wb["Percentile Analysis"]
    insert_data_into_column("A", ["Percentile"] + PERCENTILE_RANKS, 1, ws)
    insert_qs_data_into_percentile_column("B", user_qs, 1, ws, "libraries_authored")
    insert_qs_data_into_percentile_column(
        "C", user_qs, 1, ws, "library_versions_authored"
    )
    insert_qs_data_into_percentile_column(
        "F", commit_author_qs.exclude(commits_authored=0), 1, ws, "commits_authored"
    )
    insert_qs_data_into_percentile_column(
        "G", commit_author_qs.exclude(reviews_submitted=0), 1, ws, "reviews_submitted"
    )
    insert_qs_data_into_percentile_column(
        "H",
        commit_author_qs.exclude(mailing_list_count=0),
        1,
        ws,
        "mailing_list_count",
    )

    tab = Table(displayName="PercentilData", ref=f"A1:H{NUMBER_OF_PERCENTILES+1}")
    ws.add_table(tab)

    wb.save("badge_info.xlsx")
